from io import BytesIO
import pandas as pd
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_and_download_excel(tabs_data_dict):
    excel_file = BytesIO()

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        wb = writer.book  # Get the workbook object

        # Create an initial empty sheet to ensure at least one sheet exists
        if not wb.sheetnames:
            wb.create_sheet("Initial_Sheet")

        # Style settings
        if 'cell_style' not in wb.named_styles:
            cell_style = NamedStyle(name="cell_style", font=Font(size=11), alignment=Alignment(horizontal="left", vertical="center"))
            wb.add_named_style(cell_style)

        if 'title_style' not in wb.named_styles:
            title_style = NamedStyle(name="title_style", font=Font(size=12,bold=True, color="000000"), alignment=Alignment(horizontal="center", vertical="center"), fill=PatternFill(start_color="e2efe8", end_color="e2efe8", fill_type="solid"))
            wb.add_named_style(title_style)
        
        for tab, dfs_dict in tabs_data_dict.items():
            if dfs_dict:
                if tab not in wb.sheetnames:
                    wb.create_sheet(tab)
                ws = wb[tab]
                ws.sheet_state = 'visible'
                
                start_row = 1
                for key, df in dfs_dict.items():
                    df.to_excel(writer, sheet_name=tab, index=True, startrow=start_row)
                    ws = writer.sheets[tab]

                    for cell in ws[start_row]:
                        cell.style = 'cell_style'

                    for col_num, column in enumerate(ws.iter_cols(min_row=start_row, max_row=start_row + df.shape[0]+2, min_col=1, max_col=df.shape[1] + 2), 1):
                        max_length = max(len(str(cell.value)) for cell in column)
                        ws.column_dimensions[get_column_letter(col_num)].width = max_length

                        for cell in column:
                            cell.style = 'cell_style'

                    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=df.shape[1]+2)
                    title_cell = ws.cell(row=start_row, column=1, value=f'{key}')
                    title_cell.style = 'title_style'

                    start_row += df.shape[0] + 8

        # Delete the initial empty sheet if other sheets have been added
        if "Initial_Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Initial_Sheet"]

        excel_file.seek(0)

    return excel_file
